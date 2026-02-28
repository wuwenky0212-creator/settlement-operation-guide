"""Tests for export service"""
import pytest
from datetime import datetime
from io import BytesIO
import csv

from app.services.export_service import ExportService, ExportFormat, ExportResult
from app.schemas.transaction import TransactionQueryCriteria
from app.schemas.cash_flow import CashFlowQueryCriteria
from app.models.transaction import Transaction
from app.models.cash_flow import CashFlow
from app.models.enums import (
    ProductType, TransactionStatus, BackOfficeStatus,
    SettlementMethod, ConfirmationType, TransactionSource,
    Direction, CashFlowStatus
)


class TestExportService:
    """导出服务测试"""
    
    def test_export_transactions_to_csv_returns_valid_csv(self, db_session, sample_transactions):
        """测试导出交易到CSV返回有效的CSV文件"""
        # 准备测试数据
        for txn in sample_transactions:
            db_session.add(txn)
        db_session.commit()
        
        # 执行导出
        service = ExportService(db_session)
        criteria = TransactionQueryCriteria()
        result = service.export_transactions(criteria, ExportFormat.CSV)
        
        # 验证结果
        assert isinstance(result, ExportResult)
        assert result.format == ExportFormat.CSV
        assert result.record_count == len(sample_transactions)
        assert result.filename.startswith('交易汇总_')
        assert result.filename.endswith('.csv')
        assert len(result.content) > 0
        
        # 验证CSV内容
        content_str = result.content.decode('utf-8-sig')
        lines = content_str.strip().split('\n')
        assert len(lines) == len(sample_transactions) + 1  # 包含表头
        
        # 验证表头
        reader = csv.DictReader(lines)
        headers = reader.fieldnames
        assert '外部流水号' in headers
        assert '交易流水号' in headers
        assert '交易日' in headers
    
    def test_export_transactions_to_excel_returns_valid_excel(self, db_session, sample_transactions):
        """测试导出交易到Excel返回有效的Excel文件"""
        # 准备测试数据
        for txn in sample_transactions:
            db_session.add(txn)
        db_session.commit()
        
        # 执行导出
        service = ExportService(db_session)
        criteria = TransactionQueryCriteria()
        result = service.export_transactions(criteria, ExportFormat.EXCEL)
        
        # 验证结果
        assert isinstance(result, ExportResult)
        assert result.format == ExportFormat.EXCEL
        assert result.record_count == len(sample_transactions)
        assert result.filename.startswith('交易汇总_')
        assert result.filename.endswith('.excel')
        assert len(result.content) > 0
        
        # 验证Excel内容
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(result.content))
            ws = wb.active
            
            # 验证行数（包含表头）
            assert ws.max_row == len(sample_transactions) + 1
            
            # 验证表头
            headers = [cell.value for cell in ws[1]]
            assert '外部流水号' in headers
            assert '交易流水号' in headers
        except ImportError:
            pytest.skip("openpyxl not installed")
    
    def test_export_transactions_with_custom_fields(self, db_session, sample_transactions):
        """测试导出交易时指定自定义字段"""
        # 准备测试数据
        for txn in sample_transactions:
            db_session.add(txn)
        db_session.commit()
        
        # 执行导出（只导出部分字段）
        service = ExportService(db_session)
        criteria = TransactionQueryCriteria()
        custom_fields = ['external_id', 'transaction_id', 'trade_date', 'status']
        result = service.export_transactions(criteria, ExportFormat.CSV, fields=custom_fields)
        
        # 验证CSV内容
        content_str = result.content.decode('utf-8-sig')
        lines = content_str.strip().split('\n')
        reader = csv.DictReader(lines)
        headers = reader.fieldnames
        
        # 验证只包含指定字段
        assert len(headers) == len(custom_fields)
        assert '外部流水号' in headers
        assert '交易流水号' in headers
        assert '交易日' in headers
        assert '交易状态' in headers
    
    def test_export_transactions_raises_error_when_exceeds_limit(self, db_session):
        """测试导出记录数超过限制时抛出错误"""
        # 创建超过限制的记录数（通过mock）
        service = ExportService(db_session)
        service.MAX_EXPORT_RECORDS = 5  # 临时设置较小的限制
        
        # 准备测试数据（6条记录）
        for i in range(6):
            txn = Transaction(
                external_id=f'EXT-{i:03d}',
                transaction_id=f'TXN-{i:03d}',
                entry_date=datetime(2026, 2, 27, 10, 0, 0),
                trade_date=datetime(2026, 2, 27),
                value_date=datetime(2026, 2, 28),
                maturity_date=datetime(2026, 3, 28),
                account='ACC-001',
                product=ProductType.FX_SPOT,
                direction=Direction.BUY,
                underlying='USD/CNY',
                counterparty='Bank A',
                status=TransactionStatus.EFFECTIVE,
                back_office_status=BackOfficeStatus.CONFIRMED,
                settlement_method=SettlementMethod.GROSS,
                confirmation_type=ConfirmationType.SWIFT,
                nature='Normal',
                source=TransactionSource.GIT,
                operating_institution='BOCHK',
                trader='Trader A',
                version=1,
                last_modified_by='system'
            )
            db_session.add(txn)
        db_session.commit()
        
        # 执行导出并验证错误
        criteria = TransactionQueryCriteria()
        with pytest.raises(ValueError, match='EXPORT_LIMIT_EXCEEDED'):
            service.export_transactions(criteria, ExportFormat.CSV)
    
    def test_export_transactions_with_query_criteria(self, db_session, sample_transactions):
        """测试使用查询条件导出交易"""
        # 准备测试数据
        for txn in sample_transactions:
            db_session.add(txn)
        db_session.commit()
        
        # 执行导出（只导出特定状态的交易）
        service = ExportService(db_session)
        criteria = TransactionQueryCriteria(status=TransactionStatus.EFFECTIVE)
        result = service.export_transactions(criteria, ExportFormat.CSV)
        
        # 验证结果
        effective_count = sum(1 for txn in sample_transactions if txn.status == TransactionStatus.EFFECTIVE)
        assert result.record_count == effective_count
    
    def test_export_cash_flows_to_csv_returns_valid_csv(self, db_session, sample_cash_flows):
        """测试导出现金流到CSV返回有效的CSV文件"""
        # 准备测试数据
        for cf in sample_cash_flows:
            db_session.add(cf)
        db_session.commit()
        
        # 执行导出
        service = ExportService(db_session)
        criteria = CashFlowQueryCriteria()
        result = service.export_cash_flows(criteria, ExportFormat.CSV)
        
        # 验证结果
        assert isinstance(result, ExportResult)
        assert result.format == ExportFormat.CSV
        assert result.record_count == len(sample_cash_flows)
        assert result.filename.startswith('现金流_')
        assert result.filename.endswith('.csv')
        assert len(result.content) > 0
        
        # 验证CSV内容
        content_str = result.content.decode('utf-8-sig')
        lines = content_str.strip().split('\n')
        assert len(lines) == len(sample_cash_flows) + 1  # 包含表头
        
        # 验证表头
        reader = csv.DictReader(lines)
        headers = reader.fieldnames
        assert '现金流内部ID' in headers
        assert '交易流水号' in headers
        assert '币种' in headers
    
    def test_export_cash_flows_to_excel_returns_valid_excel(self, db_session, sample_cash_flows):
        """测试导出现金流到Excel返回有效的Excel文件"""
        # 准备测试数据
        for cf in sample_cash_flows:
            db_session.add(cf)
        db_session.commit()
        
        # 执行导出
        service = ExportService(db_session)
        criteria = CashFlowQueryCriteria()
        result = service.export_cash_flows(criteria, ExportFormat.EXCEL)
        
        # 验证结果
        assert isinstance(result, ExportResult)
        assert result.format == ExportFormat.EXCEL
        assert result.record_count == len(sample_cash_flows)
        assert result.filename.startswith('现金流_')
        assert result.filename.endswith('.excel')
        assert len(result.content) > 0
        
        # 验证Excel内容
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(result.content))
            ws = wb.active
            
            # 验证行数（包含表头）
            assert ws.max_row == len(sample_cash_flows) + 1
            
            # 验证表头
            headers = [cell.value for cell in ws[1]]
            assert '现金流内部ID' in headers
            assert '交易流水号' in headers
        except ImportError:
            pytest.skip("openpyxl not installed")
    
    def test_export_cash_flows_raises_error_when_exceeds_limit(self, db_session, sample_transaction):
        """测试导出现金流记录数超过限制时抛出错误"""
        # 添加交易
        db_session.add(sample_transaction)
        db_session.commit()
        
        # 创建超过限制的记录数
        service = ExportService(db_session)
        service.MAX_EXPORT_RECORDS = 3  # 临时设置较小的限制
        
        # 准备测试数据（4条记录）
        for i in range(4):
            cf = CashFlow(
                cash_flow_id=f'CF-{i:03d}',
                transaction_id=sample_transaction.transaction_id,
                direction=Direction.RECEIVE,
                currency='USD',
                amount=10000.0 + i * 1000,
                payment_date=datetime(2026, 2, 28),
                account_number='1234567890',
                account_name='Test Account',
                bank_name='Test Bank',
                bank_code='TEST001',
                settlement_method=SettlementMethod.GROSS,
                current_status=CashFlowStatus.PENDING_AML,
                progress_percentage=0,
                version=1,
                last_modified_date=datetime.now()
            )
            db_session.add(cf)
        db_session.commit()
        
        # 执行导出并验证错误
        criteria = CashFlowQueryCriteria()
        with pytest.raises(ValueError, match='EXPORT_LIMIT_EXCEEDED'):
            service.export_cash_flows(criteria, ExportFormat.CSV)
    
    def test_export_empty_result_returns_valid_file(self, db_session):
        """测试导出空结果返回有效文件"""
        # 执行导出（没有数据）
        service = ExportService(db_session)
        criteria = TransactionQueryCriteria()
        result = service.export_transactions(criteria, ExportFormat.CSV)
        
        # 验证结果
        assert result.record_count == 0
        assert len(result.content) > 0  # 至少包含表头
        
        # 验证CSV内容
        content_str = result.content.decode('utf-8-sig')
        lines = content_str.strip().split('\n')
        assert len(lines) == 1  # 只有表头
    
    def test_export_handles_special_characters_in_data(self, db_session):
        """测试导出处理数据中的特殊字符"""
        # 准备包含特殊字符的测试数据
        txn = Transaction(
            external_id='EXT-001',
            transaction_id='TXN-001',
            entry_date=datetime(2026, 2, 27, 10, 0, 0),
            trade_date=datetime(2026, 2, 27),
            value_date=datetime(2026, 2, 28),
            maturity_date=datetime(2026, 3, 28),
            account='ACC-001',
            product=ProductType.FX_SPOT,
            direction=Direction.BUY,
            underlying='USD/CNY',
            counterparty='Bank "A" & Co., Ltd.',  # 包含特殊字符
            status=TransactionStatus.EFFECTIVE,
            back_office_status=BackOfficeStatus.CONFIRMED,
            settlement_method=SettlementMethod.GROSS,
            confirmation_type=ConfirmationType.SWIFT,
            nature='Normal, Special',  # 包含逗号
            source=TransactionSource.GIT,
            operating_institution='BOCHK',
            trader='Trader A',
            version=1,
            last_modified_by='system'
        )
        db_session.add(txn)
        db_session.commit()
        
        # 执行导出
        service = ExportService(db_session)
        criteria = TransactionQueryCriteria()
        result = service.export_transactions(criteria, ExportFormat.CSV)
        
        # 验证结果
        assert result.record_count == 1
        content_str = result.content.decode('utf-8-sig')
        assert 'Bank "A" & Co., Ltd.' in content_str or 'Bank ""A"" & Co., Ltd.' in content_str
    
    def test_export_result_contains_correct_metadata(self, db_session, sample_transactions):
        """测试导出结果包含正确的元数据"""
        # 准备测试数据
        for txn in sample_transactions:
            db_session.add(txn)
        db_session.commit()
        
        # 执行导出
        service = ExportService(db_session)
        criteria = TransactionQueryCriteria()
        before_export = datetime.now()
        result = service.export_transactions(criteria, ExportFormat.CSV)
        after_export = datetime.now()
        
        # 验证元数据
        assert result.record_count == len(sample_transactions)
        assert result.format == ExportFormat.CSV
        assert before_export <= result.export_time <= after_export
        assert result.filename.startswith('交易汇总_')
        assert result.filename.endswith('.csv')


@pytest.fixture
def sample_transaction(db_session):
    """创建示例交易"""
    txn = Transaction(
        external_id='EXT-001',
        transaction_id='TXN-001',
        entry_date=datetime(2026, 2, 27, 10, 0, 0),
        trade_date=datetime(2026, 2, 27),
        value_date=datetime(2026, 2, 28),
        maturity_date=datetime(2026, 3, 28),
        account='ACC-001',
        product=ProductType.FX_SPOT,
        direction=Direction.BUY,
        underlying='USD/CNY',
        counterparty='Bank A',
        status=TransactionStatus.EFFECTIVE,
        back_office_status=BackOfficeStatus.CONFIRMED,
        settlement_method=SettlementMethod.GROSS,
        confirmation_type=ConfirmationType.SWIFT,
        nature='Normal',
        source=TransactionSource.GIT,
        operating_institution='BOCHK',
        trader='Trader A',
        version=1,
        last_modified_by='system'
    )
    return txn


@pytest.fixture
def sample_transactions():
    """创建多个示例交易"""
    transactions = []
    for i in range(5):
        txn = Transaction(
            external_id=f'EXT-{i:03d}',
            transaction_id=f'TXN-{i:03d}',
            entry_date=datetime(2026, 2, 27, 10, 0, 0),
            trade_date=datetime(2026, 2, 27 - i),
            value_date=datetime(2026, 2, 28),
            maturity_date=datetime(2026, 3, 28),
            account=f'ACC-{i:03d}',
            product=ProductType.FX_SPOT if i % 2 == 0 else ProductType.FX_FORWARD,
            direction=Direction.BUY if i % 2 == 0 else Direction.SELL,
            underlying='USD/CNY',
            counterparty=f'Bank {chr(65 + i)}',
            status=TransactionStatus.EFFECTIVE if i % 2 == 0 else TransactionStatus.MATURED,
            back_office_status=BackOfficeStatus.CONFIRMED,
            settlement_method=SettlementMethod.GROSS,
            confirmation_type=ConfirmationType.SWIFT,
            nature='Normal',
            source=TransactionSource.GIT,
            operating_institution='BOCHK',
            trader=f'Trader {chr(65 + i)}',
            version=1,
            last_modified_by='system'
        )
        transactions.append(txn)
    return transactions


@pytest.fixture
def sample_cash_flows(sample_transaction, db_session):
    """创建多个示例现金流"""
    # 先添加交易
    db_session.add(sample_transaction)
    db_session.commit()
    
    cash_flows = []
    for i in range(3):
        cf = CashFlow(
            cash_flow_id=f'CF-{i:03d}',
            transaction_id=sample_transaction.transaction_id,
            direction=Direction.RECEIVE if i % 2 == 0 else Direction.PAY,
            currency='USD' if i % 2 == 0 else 'CNY',
            amount=10000.0 + i * 1000,
            payment_date=datetime(2026, 2, 28),
            account_number=f'123456789{i}',
            account_name=f'Test Account {i}',
            bank_name=f'Test Bank {i}',
            bank_code=f'TEST00{i}',
            settlement_method=SettlementMethod.GROSS,
            current_status=CashFlowStatus.PENDING_AML if i % 2 == 0 else CashFlowStatus.AML_APPROVED,
            progress_percentage=i * 20,
            version=1,
            last_modified_date=datetime.now()
        )
        cash_flows.append(cf)
    return cash_flows
