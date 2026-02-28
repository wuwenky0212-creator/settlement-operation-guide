"""Export service for transactions and cash flows"""
from typing import Optional, List, Generator, Any
from enum import Enum
from datetime import datetime
from io import BytesIO, StringIO
import csv
from sqlalchemy.orm import Session

from app.repositories.transaction_repository import TransactionRepository
from app.repositories.cash_flow_repository import CashFlowRepository
from app.schemas.transaction import TransactionQueryCriteria
from app.schemas.cash_flow import CashFlowQueryCriteria
from app.schemas.common import PaginationParams


class ExportFormat(str, Enum):
    """导出格式"""
    EXCEL = "excel"
    CSV = "csv"


class ExportResult:
    """导出结果"""
    def __init__(
        self,
        content: bytes,
        filename: str,
        record_count: int,
        export_time: datetime,
        format: ExportFormat
    ):
        self.content = content
        self.filename = filename
        self.record_count = record_count
        self.export_time = export_time
        self.format = format


class ExportService:
    """导出服务"""
    
    # 导出记录数上限
    MAX_EXPORT_RECORDS = 10000
    
    # 流式处理批次大小（不能超过PaginationParams的page_size限制）
    BATCH_SIZE = 100
    
    def __init__(self, db: Session):
        self.db = db
        self.transaction_repo = TransactionRepository(db)
        self.cash_flow_repo = CashFlowRepository(db)
    
    def export_transactions(
        self,
        criteria: TransactionQueryCriteria,
        format: ExportFormat,
        fields: Optional[List[str]] = None
    ) -> ExportResult:
        """
        导出交易列表
        
        Args:
            criteria: 查询条件
            format: 导出格式
            fields: 要导出的字段列表，None表示导出所有字段
            
        Returns:
            ExportResult: 导出结果
            
        Raises:
            ValueError: 当导出记录数超过限制时
        """
        # 检查记录总数
        total_count = self._count_transactions(criteria)
        
        if total_count > self.MAX_EXPORT_RECORDS:
            raise ValueError(
                f"EXPORT_LIMIT_EXCEEDED: 导出记录数({total_count})超过限制({self.MAX_EXPORT_RECORDS})，请缩小查询范围"
            )
        
        # 定义默认字段
        default_fields = [
            'external_id', 'transaction_id', 'entry_date', 'trade_date',
            'value_date', 'maturity_date', 'account', 'product', 'direction',
            'underlying', 'counterparty', 'status', 'back_office_status',
            'settlement_method', 'confirmation_number', 'confirmation_type',
            'confirmation_match_type', 'nature', 'source', 'latest_event_type',
            'operating_institution', 'trader'
        ]
        
        export_fields = fields if fields else default_fields
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"交易汇总_{timestamp}.{format.value}"
        
        # 根据格式导出
        if format == ExportFormat.EXCEL:
            content = self._export_transactions_to_excel(criteria, export_fields)
        else:  # CSV
            content = self._export_transactions_to_csv(criteria, export_fields)
        
        return ExportResult(
            content=content,
            filename=filename,
            record_count=total_count,
            export_time=datetime.now(),
            format=format
        )
    
    def export_cash_flows(
        self,
        criteria: CashFlowQueryCriteria,
        format: ExportFormat
    ) -> ExportResult:
        """
        导出现金流列表
        
        Args:
            criteria: 查询条件
            format: 导出格式
            
        Returns:
            ExportResult: 导出结果
            
        Raises:
            ValueError: 当导出记录数超过限制时
        """
        # 检查记录总数
        total_count = self._count_cash_flows(criteria)
        
        if total_count > self.MAX_EXPORT_RECORDS:
            raise ValueError(
                f"EXPORT_LIMIT_EXCEEDED: 导出记录数({total_count})超过限制({self.MAX_EXPORT_RECORDS})，请缩小查询范围"
            )
        
        # 定义导出字段
        export_fields = [
            'cash_flow_id', 'transaction_id', 'direction', 'currency',
            'amount', 'payment_date', 'account_number', 'account_name',
            'bank_name', 'bank_code', 'settlement_method', 'current_status',
            'progress_percentage'
        ]
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"现金流_{timestamp}.{format.value}"
        
        # 根据格式导出
        if format == ExportFormat.EXCEL:
            content = self._export_cash_flows_to_excel(criteria, export_fields)
        else:  # CSV
            content = self._export_cash_flows_to_csv(criteria, export_fields)
        
        return ExportResult(
            content=content,
            filename=filename,
            record_count=total_count,
            export_time=datetime.now(),
            format=format
        )
    
    def _count_transactions(self, criteria: TransactionQueryCriteria) -> int:
        """统计交易记录数"""
        pagination = PaginationParams(page=1, page_size=1)
        _, total_count = self.transaction_repo.find_by_criteria(criteria, pagination)
        return total_count
    
    def _count_cash_flows(self, criteria: CashFlowQueryCriteria) -> int:
        """统计现金流记录数"""
        pagination = PaginationParams(page=1, page_size=1)
        _, total_count = self.cash_flow_repo.find_by_criteria(criteria, pagination)
        return total_count
    
    def _stream_transactions(
        self,
        criteria: TransactionQueryCriteria
    ) -> Generator[Any, None, None]:
        """
        流式读取交易记录
        
        Args:
            criteria: 查询条件
            
        Yields:
            Transaction: 交易记录
        """
        page = 1
        while True:
            pagination = PaginationParams(
                page=page,
                page_size=self.BATCH_SIZE
            )
            
            transactions, _ = self.transaction_repo.find_by_criteria(
                criteria, pagination
            )
            
            if not transactions:
                break
            
            for transaction in transactions:
                yield transaction
            
            page += 1
    
    def _stream_cash_flows(
        self,
        criteria: CashFlowQueryCriteria
    ) -> Generator[Any, None, None]:
        """
        流式读取现金流记录
        
        Args:
            criteria: 查询条件
            
        Yields:
            CashFlow: 现金流记录
        """
        page = 1
        while True:
            pagination = PaginationParams(
                page=page,
                page_size=self.BATCH_SIZE
            )
            
            cash_flows, _ = self.cash_flow_repo.find_by_criteria(
                criteria, pagination
            )
            
            if not cash_flows:
                break
            
            for cash_flow in cash_flows:
                yield cash_flow
            
            page += 1
    
    def _export_transactions_to_csv(
        self,
        criteria: TransactionQueryCriteria,
        fields: List[str]
    ) -> bytes:
        """
        导出交易到CSV
        
        Args:
            criteria: 查询条件
            fields: 导出字段列表
            
        Returns:
            bytes: CSV文件内容
        """
        output = StringIO()
        
        # 字段名映射（中文表头）
        field_names = {
            'external_id': '外部流水号',
            'transaction_id': '交易流水号',
            'entry_date': '录入日',
            'trade_date': '交易日',
            'value_date': '起息日',
            'maturity_date': '到期日',
            'account': '账户',
            'product': '产品',
            'direction': '买卖方向',
            'underlying': '标的物',
            'counterparty': '交易对手',
            'status': '交易状态',
            'back_office_status': '后线处理状态',
            'settlement_method': '清算方式',
            'confirmation_number': '证实编号',
            'confirmation_type': '证实方式',
            'confirmation_match_type': '证实匹配方式',
            'nature': '交易性质',
            'source': '交易来源',
            'latest_event_type': '事件类型',
            'operating_institution': '运营机构',
            'trader': '交易员'
        }
        
        # 写入表头
        writer = csv.DictWriter(
            output,
            fieldnames=[field_names.get(f, f) for f in fields],
            extrasaction='ignore'
        )
        writer.writeheader()
        
        # 流式写入数据
        for transaction in self._stream_transactions(criteria):
            row = {}
            for field in fields:
                value = getattr(transaction, field, '')
                
                # 格式化特殊类型
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(value, 'value'):  # Enum类型
                    value = value.value
                elif value is None:
                    value = ''
                
                row[field_names.get(field, field)] = value
            
            writer.writerow(row)
        
        # 转换为bytes（UTF-8 BOM for Excel compatibility）
        content = output.getvalue()
        return '\ufeff'.encode('utf-8') + content.encode('utf-8')
    
    def _export_transactions_to_excel(
        self,
        criteria: TransactionQueryCriteria,
        fields: List[str]
    ) -> bytes:
        """
        导出交易到Excel
        
        Args:
            criteria: 查询条件
            fields: 导出字段列表
            
        Returns:
            bytes: Excel文件内容
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "交易汇总"
        
        # 字段名映射（中文表头）
        field_names = {
            'external_id': '外部流水号',
            'transaction_id': '交易流水号',
            'entry_date': '录入日',
            'trade_date': '交易日',
            'value_date': '起息日',
            'maturity_date': '到期日',
            'account': '账户',
            'product': '产品',
            'direction': '买卖方向',
            'underlying': '标的物',
            'counterparty': '交易对手',
            'status': '交易状态',
            'back_office_status': '后线处理状态',
            'settlement_method': '清算方式',
            'confirmation_number': '证实编号',
            'confirmation_type': '证实方式',
            'confirmation_match_type': '证实匹配方式',
            'nature': '交易性质',
            'source': '交易来源',
            'latest_event_type': '事件类型',
            'operating_institution': '运营机构',
            'trader': '交易员'
        }
        
        # 写入表头
        headers = [field_names.get(f, f) for f in fields]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 流式写入数据
        for transaction in self._stream_transactions(criteria):
            row = []
            for field in fields:
                value = getattr(transaction, field, '')
                
                # 格式化特殊类型
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(value, 'value'):  # Enum类型
                    value = value.value
                elif value is None:
                    value = ''
                
                row.append(value)
            
            ws.append(row)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
    
    def _export_cash_flows_to_csv(
        self,
        criteria: CashFlowQueryCriteria,
        fields: List[str]
    ) -> bytes:
        """
        导出现金流到CSV
        
        Args:
            criteria: 查询条件
            fields: 导出字段列表
            
        Returns:
            bytes: CSV文件内容
        """
        output = StringIO()
        
        # 字段名映射（中文表头）
        field_names = {
            'cash_flow_id': '现金流内部ID',
            'transaction_id': '交易流水号',
            'direction': '方向',
            'currency': '币种',
            'amount': '金额',
            'payment_date': '收付日期',
            'account_number': '账号',
            'account_name': '户名',
            'bank_name': '开户行',
            'bank_code': '开户行号',
            'settlement_method': '结算方式',
            'current_status': '当前状态',
            'progress_percentage': '进度百分比'
        }
        
        # 写入表头
        writer = csv.DictWriter(
            output,
            fieldnames=[field_names.get(f, f) for f in fields],
            extrasaction='ignore'
        )
        writer.writeheader()
        
        # 流式写入数据
        for cash_flow in self._stream_cash_flows(criteria):
            row = {}
            for field in fields:
                value = getattr(cash_flow, field, '')
                
                # 格式化特殊类型
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(value, 'value'):  # Enum类型
                    value = value.value
                elif value is None:
                    value = ''
                
                row[field_names.get(field, field)] = value
            
            writer.writerow(row)
        
        # 转换为bytes（UTF-8 BOM for Excel compatibility）
        content = output.getvalue()
        return '\ufeff'.encode('utf-8') + content.encode('utf-8')
    
    def _export_cash_flows_to_excel(
        self,
        criteria: CashFlowQueryCriteria,
        fields: List[str]
    ) -> bytes:
        """
        导出现金流到Excel
        
        Args:
            criteria: 查询条件
            fields: 导出字段列表
            
        Returns:
            bytes: Excel文件内容
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "现金流"
        
        # 字段名映射（中文表头）
        field_names = {
            'cash_flow_id': '现金流内部ID',
            'transaction_id': '交易流水号',
            'direction': '方向',
            'currency': '币种',
            'amount': '金额',
            'payment_date': '收付日期',
            'account_number': '账号',
            'account_name': '户名',
            'bank_name': '开户行',
            'bank_code': '开户行号',
            'settlement_method': '结算方式',
            'current_status': '当前状态',
            'progress_percentage': '进度百分比'
        }
        
        # 写入表头
        headers = [field_names.get(f, f) for f in fields]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 流式写入数据
        for cash_flow in self._stream_cash_flows(criteria):
            row = []
            for field in fields:
                value = getattr(cash_flow, field, '')
                
                # 格式化特殊类型
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(value, 'value'):  # Enum类型
                    value = value.value
                elif value is None:
                    value = ''
                
                row.append(value)
            
            ws.append(row)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
